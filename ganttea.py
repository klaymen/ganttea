#!/usr/bin/env python3
"""Generate a standalone HTML Gantt chart from an Excel input file."""

from __future__ import annotations

import argparse
import html
import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

TYPE_STYLES = {
    "Task": {"color": "#3b82f6", "label": "Task"},
    "Holiday": {"color": "#94a3b8", "label": "Holiday"},
    "Milestone": {"color": "#f59e0b", "label": "Milestone"},
}

DEFAULT_TYPE_STYLE = {"color": "#64748b", "label": "Other"}


def parse_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip().rstrip(".")
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def load_tasks(input_path: Path) -> list[dict]:
    workbook = load_workbook(input_path, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"No data found in {input_path}")

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    header_map = {name.lower(): idx for idx, name in enumerate(headers)}

    required = ["task", "type", "start date"]
    missing = [name for name in required if name not in header_map]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    tasks: list[dict] = []
    for row in rows[1:]:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        task_name = row[header_map["task"]]
        task_type = row[header_map["type"]]
        if task_name is None or str(task_name).strip() == "":
            continue

        responsible_idx = header_map.get("responsible")
        responsible = (
            str(row[responsible_idx]).strip()
            if responsible_idx is not None and row[responsible_idx] is not None
            else "Unassigned"
        )

        start = parse_date(row[header_map["start date"]])
        end_idx = header_map.get("end date")
        end = parse_date(row[end_idx]) if end_idx is not None else None

        if start is None:
            continue

        type_name = str(task_type).strip() if task_type is not None else "Task"
        type_key = type_name.lower()
        is_milestone = type_key == "milestone"

        if is_milestone:
            end = start
        elif end is None:
            end = start

        tasks.append(
            {
                "task": str(task_name).strip(),
                "type": type_name,
                "responsible": responsible or "Unassigned",
                "start": start.isoformat(),
                "end": end.isoformat(),
                "is_milestone": is_milestone,
            }
        )

    if not tasks:
        raise ValueError("No valid task rows found in the input file.")

    return tasks


def build_html(tasks: list[dict], title: str) -> str:
    starts = [task["start"] for task in tasks]
    ends = [task["end"] for task in tasks if not task["is_milestone"]]
    milestone_starts = [task["start"] for task in tasks if task["is_milestone"]]
    min_date = min(starts)
    max_date = max(ends + milestone_starts)

    milestones = [task for task in tasks if task["is_milestone"]]
    regular_tasks = [task for task in tasks if not task["is_milestone"]]

    people: dict[str, list[dict]] = {}
    for task in regular_tasks:
        people.setdefault(task["responsible"], []).append(task)

    for person_tasks in people.values():
        person_tasks.sort(key=lambda item: (item["start"], item["task"]))

    legend_items = []
    seen_types: set[str] = set()
    for task in tasks:
        task_type = task["type"]
        if task_type not in seen_types:
            seen_types.add(task_type)
            style = TYPE_STYLES.get(task_type, DEFAULT_TYPE_STYLE)
            legend_items.append({"type": task_type, "color": style["color"]})

    payload = {
        "title": title,
        "minDate": min_date,
        "maxDate": max_date,
        "people": people,
        "milestones": milestones,
        "typeStyles": TYPE_STYLES,
        "defaultTypeStyle": DEFAULT_TYPE_STYLE,
        "legend": legend_items,
    }

    data_json = json.dumps(payload)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(title)}</title>
  <style>
    :root {{
      --bg: #f1f5f9;
      --panel: #ffffff;
      --panel-muted: #f8fafc;
      --text: #0f172a;
      --muted: #64748b;
      --border: #dbe3ee;
      --grid: rgba(100, 116, 139, 0.07);
      --shadow: rgba(15, 23, 42, 0.06);
      --bar-text: #ffffff;
      --toggle-bg: #e2e8f0;
      --toggle-text: #334155;
      --sidebar-width: 280px;
      --row-height: 1.8rem;
      --header-height: 4rem;
    }}

    [data-theme="dark"] {{
      --bg: #0b1220;
      --panel: #111827;
      --panel-muted: #0f172a;
      --text: #e5eefb;
      --muted: #94a3b8;
      --border: #243044;
      --grid: rgba(148, 163, 184, 0.08);
      --shadow: rgba(0, 0, 0, 0.35);
      --bar-text: #f8fafc;
      --toggle-bg: #1e293b;
      --toggle-text: #cbd5e1;
    }}

    [data-theme="ocean"] {{
      --bg: #0d1b2a;
      --panel: #112236;
      --panel-muted: #0d1b2a;
      --text: #cce7ff;
      --muted: #5b9ec9;
      --border: #1a3a55;
      --grid: rgba(91, 158, 201, 0.08);
      --shadow: rgba(0, 0, 0, 0.4);
      --bar-text: #ffffff;
      --toggle-bg: #1a3a55;
      --toggle-text: #cce7ff;
    }}

    [data-theme="forest"] {{
      --bg: #f0f7f0;
      --panel: #ffffff;
      --panel-muted: #f4faf4;
      --text: #1a2e1a;
      --muted: #4a7c59;
      --border: #c5dfc5;
      --grid: rgba(74, 124, 89, 0.07);
      --shadow: rgba(26, 46, 26, 0.07);
      --bar-text: #ffffff;
      --toggle-bg: #dceedd;
      --toggle-text: #1a2e1a;
    }}

    [data-theme="sunset"] {{
      --bg: #1a0a0f;
      --panel: #2a1018;
      --panel-muted: #1a0a0f;
      --text: #fce8d8;
      --muted: #c47f6a;
      --border: #3d1a22;
      --grid: rgba(196, 127, 106, 0.08);
      --shadow: rgba(0, 0, 0, 0.45);
      --bar-text: #ffffff;
      --toggle-bg: #3d1a22;
      --toggle-text: #fce8d8;
    }}

    * {{ box-sizing: border-box; }}

    body {{
      margin: 0;
      font-family: Inter, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      background: var(--bg);
      color: var(--text);
      font-size: 18px;
      line-height: 1.35;
    }}

    .page {{
      max-width: 1440px;
      margin: 0 auto;
      padding: 16px;
    }}

    .header {{
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      align-items: center;
      justify-content: space-between;
      margin-bottom: 12px;
    }}

    .header-left {{
      display: flex;
      align-items: center;
      gap: 12px;
      min-width: 0;
    }}

    h1 {{
      margin: 0;
      font-size: 1.15rem;
      font-weight: 700;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }}

    .controls {{
      display: flex;
      align-items: center;
      gap: 10px;
      flex-wrap: wrap;
    }}

    .theme-toggle {{
      border: 1px solid var(--border);
      background: var(--toggle-bg);
      color: var(--toggle-text);
      border-radius: 8px;
      padding: 5px 9px;
      font-size: 0.72rem;
      font-weight: 600;
      cursor: pointer;
      transition: filter 0.15s;
    }}

    .theme-toggle:hover {{
      filter: brightness(1.08);
    }}

    .theme-toggle.active {{
      outline: 2px solid var(--muted);
      outline-offset: 1px;
    }}

    .legend {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
    }}

    .legend-item {{
      display: inline-flex;
      align-items: center;
      gap: 6px;
      font-size: 0.72rem;
      color: var(--muted);
    }}

    .legend-swatch {{
      width: 10px;
      height: 10px;
      border-radius: 2px;
      border: 1px solid rgba(127, 127, 127, 0.25);
      flex-shrink: 0;
    }}

    .legend-swatch.milestone {{
      transform: rotate(45deg);
      border-radius: 1px;
    }}

    .chart-shell {{
      background: var(--panel);
      border: 1px solid var(--border);
      border-radius: 10px;
      overflow: hidden;
      box-shadow: 0 6px 18px var(--shadow);
    }}

    .chart-scroll {{
      overflow-x: auto;
    }}

    .chart {{
      min-width: 820px;
    }}

    .timeline-header,
    .person-row,
    .task-row,
    .milestone-row {{
      display: grid;
      grid-template-columns: var(--sidebar-width) 1fr;
    }}

    .timeline-header {{
      border-bottom: 1px solid var(--border);
      background: var(--panel-muted);
    }}

    .sidebar-header,
    .timeline-labels {{
      padding: 8px 10px;
      font-size: 0.65rem;
      font-weight: 700;
      letter-spacing: 0.06em;
      text-transform: uppercase;
      color: var(--muted);
    }}

    .timeline-labels {{
      position: relative;
      height: var(--header-height);
      border-left: 1px solid var(--border);
    }}

    .timeline-separator {{
      position: absolute;
      top: 1.8rem;
      left: 0;
      right: 0;
      height: 1px;
      background: var(--border);
    }}

    .month-label {{
      position: absolute;
      top: 0.4rem;
      font-size: 0.68rem;
      font-weight: 600;
      color: var(--text);
      transform: translateX(-50%);
      white-space: nowrap;
    }}

    .tick {{
      position: absolute;
      top: 1.1rem;
      width: 1px;
      height: 0.7rem;
      background: var(--border);
    }}

    .day-label {{
      position: absolute;
      top: 2.1rem;
      font-size: 0.60rem;
      font-weight: 500;
      color: var(--muted);
      transform: translateX(-50%);
      white-space: nowrap;
    }}

    .day-tick {{
      position: absolute;
      top: 3rem;
      width: 1px;
      height: 0.8rem;
      background: var(--border);
      opacity: 0.5;
    }}

    .section-label {{
      padding: 5px 10px;
      font-size: 0.65rem;
      font-weight: 700;
      letter-spacing: 0.06em;
      text-transform: uppercase;
      color: var(--muted);
      background: var(--panel-muted);
      border-top: 1px solid var(--border);
      border-bottom: 1px solid var(--border);
    }}

    .person-group {{
      border-top: 1px solid var(--border);
    }}

    .person-row {{
      min-height: calc(var(--row-height) + 2px);
      background: var(--panel-muted);
      border-bottom: 1px solid var(--border);
    }}

    .person-name {{
      padding: 0 10px;
      font-weight: 700;
      font-size: 0.75rem;
      display: flex;
      align-items: center;
    }}

    .person-track,
    .task-track,
    .milestone-track {{
      border-left: 1px solid var(--border);
      position: relative;
      min-height: var(--row-height);
      background: repeating-linear-gradient(
        to right,
        var(--grid) 0,
        var(--grid) calc(100% / var(--week-count)),
        transparent calc(100% / var(--week-count)),
        transparent calc(200% / var(--week-count))
      );
    }}

    .person-track {{
      min-height: calc(var(--row-height) + 2px);
    }}

    .task-row,
    .milestone-row {{
      min-height: var(--row-height);
      border-bottom: 1px solid var(--border);
    }}

    .task-row:hover,
    .milestone-row:hover {{
      background: color-mix(in srgb, var(--panel-muted) 65%, transparent);
    }}

    .task-label,
    .milestone-label-cell {{
      padding: 0 10px 0 18px;
      display: flex;
      align-items: center;
      gap: 8px;
      min-width: 0;
      font-size: 0.72rem;
    }}

    .task-label .name,
    .milestone-label-cell .name {{
      font-weight: 600;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }}

    .type-badge {{
      flex-shrink: 0;
      font-size: 0.62rem;
      font-weight: 600;
      text-transform: uppercase;
      letter-spacing: 0.04em;
      color: var(--muted);
      border: 1px solid var(--border);
      border-radius: 999px;
      padding: 1px 6px;
      line-height: 1.2;
    }}

    .bar {{
      position: absolute;
      top: 0.2rem;
      bottom: 0.2rem;
      height: auto;
      border-radius: 3px;
      display: flex;
      align-items: center;
      padding: 0 6px;
      color: var(--bar-text);
      font-size: 0.65rem;
      font-weight: 600;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
      box-shadow: inset 0 1px 0 rgba(255, 255, 255, 0.12);
    }}

    .bar.holiday {{
      background-image: repeating-linear-gradient(
        -45deg,
        rgba(255, 255, 255, 0.14) 0,
        rgba(255, 255, 255, 0.14) 4px,
        transparent 4px,
        transparent 8px
      );
    }}

    .milestone-marker {{
      position: absolute;
      top: 50%;
      width: 10px;
      height: 10px;
      transform: translate(-50%, -50%) rotate(45deg);
      border-radius: 1px;
      border: 2px solid var(--panel);
      box-shadow: 0 0 0 1px rgba(127, 127, 127, 0.35);
      z-index: 2;
    }}

    .gantt-tooltip {{
      position: fixed;
      z-index: 9999;
      pointer-events: none;
      background: var(--panel);
      border: 1px solid var(--border);
      border-radius: 6px;
      padding: 7px 10px;
      box-shadow: 0 4px 16px var(--shadow);
      display: none;
      max-width: 240px;
    }}

    .gantt-tooltip-title {{
      font-size: 0.75rem;
      font-weight: 700;
      margin-bottom: 3px;
      color: var(--text);
    }}

    .gantt-tooltip-dates {{
      font-size: 0.70rem;
      color: var(--muted);
      white-space: pre-line;
      line-height: 1.5;
    }}

    .footer {{
      margin-top: 8px;
      color: var(--muted);
      font-size: 0.72rem;
    }}
  </style>
</head>
<body>
  <div class="page">
    <div class="header">
      <div class="header-left">
        <h1 id="chart-title"></h1>
      </div>
      <div class="controls">
        <div id="theme-switcher" role="group" aria-label="Color scheme" style="display:flex;gap:6px;flex-wrap:wrap;align-items:center"></div>
        <div id="font-size-control" style="display:flex;align-items:center;gap:4px">
          <button class="theme-toggle" id="fs-down" type="button" aria-label="Decrease font size">A−</button>
          <span id="fs-label" style="font-size:0.72rem;font-weight:600;color:var(--muted);min-width:28px;text-align:center"></span>
          <button class="theme-toggle" id="fs-up"   type="button" aria-label="Increase font size">A+</button>
        </div>
        <div class="legend" id="legend"></div>
      </div>
    </div>
    <div class="chart-shell">
      <div class="chart-scroll">
        <div class="chart" id="chart"></div>
      </div>
    </div>
    <div class="footer" id="range-label"></div>
  </div>

  <script>
    const DATA = {data_json};

    let tooltip = null;

    function parseDate(value) {{
      const [year, month, day] = value.split("-").map(Number);
      return new Date(year, month - 1, day);
    }}

    function dayDiff(start, end) {{
      return Math.round((end - start) / (1000 * 60 * 60 * 24));
    }}

    function addDays(date, days) {{
      const copy = new Date(date);
      copy.setDate(copy.getDate() + days);
      return copy;
    }}

    function startOfMonth(date) {{
      return new Date(date.getFullYear(), date.getMonth(), 1);
    }}

    function endOfMonth(date) {{
      return new Date(date.getFullYear(), date.getMonth() + 1, 0);
    }}

    function formatDate(date) {{
      return date.toLocaleDateString(undefined, {{
        year: "numeric",
        month: "short",
        day: "numeric",
      }});
    }}

    function getTypeStyle(typeName) {{
      return DATA.typeStyles[typeName] || DATA.defaultTypeStyle;
    }}

    const PERSON_PALETTE = [
      "#e63946", // vivid red
      "#2a9d8f", // teal
      "#f4a261", // warm orange
      "#457b9d", // steel blue
      "#9b5de5", // purple
      "#06d6a0", // mint
      "#fb5607", // deep orange
      "#3a86ff", // bright blue
      "#ff006e", // hot pink
      "#8ac926", // lime green
    ];

    function buildPersonColors(people) {{
      const map = {{}};
      Object.keys(people).forEach((person, i) => {{
        map[person] = PERSON_PALETTE[i % PERSON_PALETTE.length];
      }});
      return map;
    }}

    function isMilestone(task) {{
      return task.is_milestone === true || String(task.type || "").toLowerCase() === "milestone";
    }}

    function buildTimeline(minDate, maxDate, width) {{
      const totalDays = Math.max(dayDiff(minDate, maxDate), 1);
      const weekCount = Math.max(Math.ceil(totalDays / 7), 1);

      const labels = document.createElement("div");
      labels.className = "timeline-labels";
      labels.style.setProperty("--week-count", weekCount);

      const separator = document.createElement("div");
      separator.className = "timeline-separator";
      labels.appendChild(separator);

      const pxPerDay = width / totalDays;
      let dayInterval = 0;
      if (pxPerDay >= 8) dayInterval = 7;
      else if (pxPerDay >= 4) dayInterval = 14;
      else if (pxPerDay >= 2) dayInterval = 28;

      // Collect month-start pixel positions so day labels can avoid them
      const monthStartPx = [];
      let cursor = startOfMonth(minDate);
      const endLimit = endOfMonth(maxDate);

      while (cursor <= endLimit) {{
        const offsetDays = dayDiff(minDate, cursor);
        const left = (offsetDays / totalDays) * width;

        // Advance cursor first so we can use continue cleanly
        const thisMonth = cursor;
        cursor = new Date(cursor.getFullYear(), cursor.getMonth() + 1, 1);

        // Skip months that start before the visible range
        if (offsetDays < 0) continue;

        monthStartPx.push(left);

        const label = document.createElement("div");
        label.className = "month-label";
        label.style.left = `${{left}}px`;
        // If the label sits at the very left edge, anchor it there instead of centering
        if (left < 2) label.style.transform = "none";
        label.textContent = thisMonth.toLocaleDateString(undefined, {{
          month: "short",
          year: "2-digit",
        }});
        labels.appendChild(label);

        const tick = document.createElement("div");
        tick.className = "tick";
        tick.style.left = `${{left}}px`;
        labels.appendChild(tick);
      }}

      if (dayInterval > 0) {{
        const clearance = pxPerDay * dayInterval * 0.55;
        for (let d = 0; ; d += dayInterval) {{
          const date = addDays(minDate, d);
          if (date > maxDate) break;
          const left = (d / totalDays) * width;

          // Skip day labels that sit too close to a month-start tick
          const tooClose = monthStartPx.some(mx => Math.abs(left - mx) < clearance);
          if (tooClose) continue;

          const dayLabel = document.createElement("div");
          dayLabel.className = "day-label";
          dayLabel.style.left = `${{left}}px`;
          dayLabel.textContent = date.getDate();
          labels.appendChild(dayLabel);

          const dayTick = document.createElement("div");
          dayTick.className = "day-tick";
          dayTick.style.left = `${{left}}px`;
          labels.appendChild(dayTick);
        }}
      }}

      return {{ labels, totalDays, weekCount }};
    }}

    function positionForDate(date, minDate, totalDays, trackWidth) {{
      const offsetDays = dayDiff(minDate, date);
      return (offsetDays / totalDays) * trackWidth;
    }}

    function createTrack(className, trackWidth, weekCount) {{
      const track = document.createElement("div");
      track.className = className;
      track.style.minWidth = `${{trackWidth}}px`;
      track.style.setProperty("--week-count", weekCount);
      return track;
    }}

    function renderBar(task, track, minDate, timeline, trackWidth, personColor) {{
      const style = getTypeStyle(task.type);
      const start = parseDate(task.start);
      const end = parseDate(task.end);
      const typeKey = String(task.type || "").toLowerCase();

      if (isMilestone(task)) {{
        return;
      }}

      const left = positionForDate(start, minDate, timeline.totalDays, trackWidth);
      const inclusiveEnd = addDays(end, 1);
      const right = positionForDate(inclusiveEnd, minDate, timeline.totalDays, trackWidth);
      const width = Math.max(right - left, 4);

      const bar = document.createElement("div");
      bar.className = "bar" + (typeKey === "holiday" ? " holiday" : "");
      bar.style.left = `${{left}}px`;
      bar.style.width = `${{width}}px`;
      bar.style.backgroundColor = (typeKey === "task" && personColor) ? personColor : style.color;
      bar.addEventListener("mouseenter", (e) => tooltip.show(e, task.task, `Start: ${{formatDate(start)}}\nEnd:   ${{formatDate(end)}}`));
      bar.addEventListener("mousemove", (e) => tooltip.move(e));
      bar.addEventListener("mouseleave", () => tooltip.hide());
      if (width >= 36) {{
        bar.textContent = task.task;
      }}
      track.appendChild(bar);
    }}

    function renderMilestone(task, track, minDate, timeline, trackWidth) {{
      if (!isMilestone(task)) {{
        return;
      }}

      const style = getTypeStyle(task.type);
      const start = parseDate(task.start);
      const left = positionForDate(start, minDate, timeline.totalDays, trackWidth);

      const marker = document.createElement("div");
      marker.className = "milestone-marker";
      marker.style.left = `${{left}}px`;
      marker.style.background = style.color;
      marker.addEventListener("mouseenter", (e) => tooltip.show(e, task.task, `Date: ${{formatDate(start)}}`));
      marker.addEventListener("mousemove", (e) => tooltip.move(e));
      marker.addEventListener("mouseleave", () => tooltip.hide());
      track.appendChild(marker);
    }}

    function renderPersonGroup(person, tasks, minDate, timeline, trackWidth, personColor) {{
      const group = document.createElement("div");
      group.className = "person-group";

      const personRow = document.createElement("div");
      personRow.className = "person-row";

      const personName = document.createElement("div");
      personName.className = "person-name";
      personName.textContent = person;
      if (personColor) {{
        personName.style.color = personColor;
        personName.style.borderLeft = `3px solid ${{personColor}}`;
        personName.style.paddingLeft = `7px`;
      }}

      const personTrack = createTrack("person-track", trackWidth, timeline.weekCount);
      personRow.appendChild(personName);
      personRow.appendChild(personTrack);
      group.appendChild(personRow);

      tasks.forEach((task) => {{
        if (isMilestone(task)) {{
          return;
        }}

        const row = document.createElement("div");
        row.className = "task-row";

        const label = document.createElement("div");
        label.className = "task-label";

        const name = document.createElement("span");
        name.className = "name";
        name.textContent = task.task;

        const badge = document.createElement("span");
        badge.className = "type-badge";
        badge.textContent = task.type;

        label.appendChild(name);
        label.appendChild(badge);

        const track = createTrack("task-track", trackWidth, timeline.weekCount);
        renderBar(task, track, minDate, timeline, trackWidth, personColor);

        row.appendChild(label);
        row.appendChild(track);
        group.appendChild(row);
      }});

      return group;
    }}

    function renderMilestonesSection(milestones, minDate, timeline, trackWidth) {{
      if (!milestones.length) {{
        return null;
      }}

      const section = document.createElement("div");
      section.className = "person-group";

      const heading = document.createElement("div");
      heading.className = "section-label";
      heading.textContent = "Milestones";
      section.appendChild(heading);

      milestones.forEach((task) => {{
        const row = document.createElement("div");
        row.className = "milestone-row";

        const label = document.createElement("div");
        label.className = "milestone-label-cell";

        const name = document.createElement("span");
        name.className = "name";
        name.textContent = task.task;

        const badge = document.createElement("span");
        badge.className = "type-badge";
        badge.textContent = task.responsible === "Unassigned" ? "Global" : task.responsible;

        label.appendChild(name);
        label.appendChild(badge);

        const track = createTrack("milestone-track", trackWidth, timeline.weekCount);
        renderMilestone(task, track, minDate, timeline, trackWidth);

        row.appendChild(label);
        row.appendChild(track);
        section.appendChild(row);
      }});

      return section;
    }}

    function initTheme() {{
      const root = document.documentElement;
      const switcher = document.getElementById("theme-switcher");
      const stored = localStorage.getItem("ganttea-theme");
      const prefersDark = window.matchMedia("(prefers-color-scheme: dark)").matches;
      const defaultTheme = stored || (prefersDark ? "dark" : "light");

      const schemes = [
        {{ id: "light",  label: "Light"  }},
        {{ id: "dark",   label: "Dark"   }},
        {{ id: "ocean",  label: "Ocean"  }},
        {{ id: "forest", label: "Forest" }},
        {{ id: "sunset", label: "Sunset" }},
      ];

      function applyTheme(id) {{
        root.setAttribute("data-theme", id);
        localStorage.setItem("ganttea-theme", id);
        switcher.querySelectorAll(".theme-toggle").forEach(btn => {{
          btn.classList.toggle("active", btn.dataset.theme === id);
        }});
      }}

      schemes.forEach(scheme => {{
        const btn = document.createElement("button");
        btn.type = "button";
        btn.className = "theme-toggle";
        btn.dataset.theme = scheme.id;
        btn.textContent = scheme.label;
        btn.addEventListener("click", () => applyTheme(scheme.id));
        switcher.appendChild(btn);
      }});

      applyTheme(defaultTheme);
    }}

    function initFontSize() {{
      const SIZES = [10, 11, 12, 13, 14, 16, 18, 20, 22, 24, 28, 32];
      const DEFAULT = 18;
      const KEY = "ganttea-fontsize";
      const root = document.documentElement;
      const label = document.getElementById("fs-label");

      let current = parseInt(localStorage.getItem(KEY), 10);
      if (!SIZES.includes(current)) current = DEFAULT;

      function apply(size) {{
        current = size;
        root.style.fontSize = size + "px";
        label.textContent = size + "px";
        localStorage.setItem(KEY, size);
        document.getElementById("fs-down").disabled = size <= SIZES[0];
        document.getElementById("fs-up").disabled  = size >= SIZES[SIZES.length - 1];
      }}

      document.getElementById("fs-down").addEventListener("click", () => {{
        const idx = SIZES.indexOf(current);
        if (idx > 0) apply(SIZES[idx - 1]);
      }});
      document.getElementById("fs-up").addEventListener("click", () => {{
        const idx = SIZES.indexOf(current);
        if (idx < SIZES.length - 1) apply(SIZES[idx + 1]);
      }});

      apply(current);
    }}

    function setupTooltip() {{
      const el = document.createElement("div");
      el.className = "gantt-tooltip";
      el.innerHTML = '<div class="gantt-tooltip-title"></div><div class="gantt-tooltip-dates"></div>';
      document.body.appendChild(el);

      function show(e, title, dateText) {{
        el.querySelector(".gantt-tooltip-title").textContent = title;
        el.querySelector(".gantt-tooltip-dates").textContent = dateText;
        el.style.display = "block";
        move(e);
      }}

      function move(e) {{
        const gap = 14;
        let x = e.clientX + gap;
        let y = e.clientY - el.offsetHeight - gap;
        if (x + el.offsetWidth > window.innerWidth - 8) x = e.clientX - el.offsetWidth - gap;
        if (y < 8) y = e.clientY + gap;
        el.style.left = `${{x}}px`;
        el.style.top = `${{y}}px`;
      }}

      function hide() {{
        el.style.display = "none";
      }}

      return {{ show, move, hide }};
    }}

    function render() {{
      tooltip = setupTooltip();
      const minDate = parseDate(DATA.minDate);
      const maxDate = parseDate(DATA.maxDate);
      const chart = document.getElementById("chart");
      const trackWidth = Math.max(760, dayDiff(minDate, maxDate) * 10);

      document.getElementById("chart-title").textContent = DATA.title;
      document.getElementById("range-label").textContent =
        `Timeline: ${{formatDate(minDate)}} – ${{formatDate(maxDate)}}`;

      const legend = document.getElementById("legend");
      DATA.legend.forEach((item) => {{
        const entry = document.createElement("div");
        entry.className = "legend-item";

        const swatch = document.createElement("span");
        swatch.className = "legend-swatch" + (item.type.toLowerCase() === "milestone" ? " milestone" : "");
        swatch.style.background = item.color;

        const text = document.createElement("span");
        text.textContent = item.type;

        entry.appendChild(swatch);
        entry.appendChild(text);
        legend.appendChild(entry);
      }});

      const header = document.createElement("div");
      header.className = "timeline-header";

      const sidebarHeader = document.createElement("div");
      sidebarHeader.className = "sidebar-header";
      sidebarHeader.textContent = "Task";

      const timeline = buildTimeline(minDate, maxDate, trackWidth);
      header.appendChild(sidebarHeader);
      header.appendChild(timeline.labels);
      chart.appendChild(header);

      const personColors = buildPersonColors(DATA.people);
      Object.entries(DATA.people).forEach(([person, tasks]) => {{
        chart.appendChild(renderPersonGroup(person, tasks, minDate, timeline, trackWidth, personColors[person]));
      }});

      const milestonesSection = renderMilestonesSection(DATA.milestones || [], minDate, timeline, trackWidth);
      if (milestonesSection) {{
        chart.appendChild(milestonesSection);
      }}

      initTheme();
      initFontSize();
    }}

    render();
  </script>
</body>
</html>
"""


def generate_gantt(input_path: Path, output_path: Path, title: str | None = None) -> None:
    tasks = load_tasks(input_path)
    chart_title = title or f"Gantt Chart — {input_path.stem}"
    html_content = build_html(tasks, chart_title)
    output_path.write_text(html_content, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate a standalone HTML Gantt chart.")
    parser.add_argument(
        "input",
        nargs="?",
        default="test_input.xlsx",
        help="Input Excel file (.xlsx). Default: test_input.xlsx",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="gantt_chart.html",
        help="Output HTML file path. Default: gantt_chart.html",
    )
    parser.add_argument(
        "-t",
        "--title",
        default=None,
        help="Optional chart title.",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")

    generate_gantt(input_path, output_path, args.title)
    print(f"Gantt chart written to {output_path.resolve()}")


if __name__ == "__main__":
    main()
